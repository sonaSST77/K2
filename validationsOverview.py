# validationsOverview.py
# Připojení k DB, načtení dat z MSA_VAL_DESIGN, export do Excelu

import oracledb
import pandas as pd
import os
from struktura_Dat.db_connect import get_db_connection

def main():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        query = "SELECT * FROM K2_MIGUSER1.MSA_VAL_DESIGN ORDER BY VAL_ID"
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        data = cursor.fetchall()
        df = pd.DataFrame(data, columns=columns)
        output_dir = "vystupy"
        os.makedirs(output_dir, exist_ok=True)
        now_str = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"valDesign_{now_str}.xlsx")
        # ...existing code...
        # Druhý list: vybrané sloupce a číslo z rowcount ve FINAL_QUERY
        cols_needed = ["VAL_ID", "NAME", "TABLE_NAME", "COLUMN_NAME", "SEVERITY", "STATUS", "FINAL_QUERY"]
        df2 = df[cols_needed].copy() if all(col in df.columns for col in cols_needed) else pd.DataFrame()
        # Načíst kategorie z DB
        kategory_dict = {}
        kategory_list = []
        try:
            cursor_kat = connection.cursor()
            cursor_kat.execute("SELECT VAL_ID, KATEGORY FROM SO081267.TMP_SST_VALIDATION_KATEGORY")
            for val_id, kategory in cursor_kat.fetchall():
                kategory_dict[val_id] = kategory
                kategory_list.append(kategory)
            cursor_kat.close()
        except Exception as e:
            print("Chyba při načítání kategorií z DB:", e)
        if not df2.empty:
            import re
            def extract_rowcount(val):
                if pd.isnull(val):
                    return None
                match = re.search(r"rowcount\s*=\s*(\d+)", str(val))
                return int(match.group(1)) if match else None
            def extract_sysdate(val):
                if pd.isnull(val):
                    return None
                import re
                match = re.search(r"sysdate\s*[=:]?\s*([0-9]{2})\.([0-9]{2})\.([0-9]{4})", str(val))
                if match:
                    return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
                match2 = re.search(r"([0-9]{4}-[0-9]{2}-[0-9]{2})", str(val))
                return match2.group(1) if match2 else None
            df2["DATUM_KONTROLY"] = df2["FINAL_QUERY"].apply(extract_sysdate)
            df2["ROWCOUNT"] = df2["FINAL_QUERY"].apply(extract_rowcount)
            df2 = df2.drop(columns=["FINAL_QUERY"])
            # Přidat sloupec Kategory podle DB, jinak 'nedefinováno'
            df2.insert(0, "Kategory", df2["VAL_ID"].map(lambda x: kategory_dict.get(x, "nedefinováno")))
            # Přesunout sloupec DATUM_KONTROLY před ROWCOUNT (vždy)
            cols = df2.columns.tolist()
            if "DATUM_KONTROLY" in cols and "ROWCOUNT" in cols:
                cols.remove("DATUM_KONTROLY")
                cols.insert(cols.index("ROWCOUNT"), "DATUM_KONTROLY")
                df2 = df2[cols]
        # List kategory: načíst hodnoty z tabulky SO081267.TMP_SST_VALIDATION_KATEGORY_CIS
        try:
            cursor_cis = connection.cursor()
            cursor_cis.execute("SELECT KATEGORY, ZODPOVEDNA_OSOBA, KONTAKT, TERMIN FROM SO081267.TMP_SST_VALIDATION_KATEGORY_CIS")
            data_cis = cursor_cis.fetchall()
            df_kategory = pd.DataFrame(data_cis, columns=["KATEGORY", "Zodpovědná osoba", "kontakt", "Termín"])
            cursor_cis.close()
        except Exception as e:
            print("Chyba při načítání kategory cis z DB:", e)
            df_kategory = pd.DataFrame({
                "KATEGORIE": ["nedefinováno"],
                "Zodpovědná osoba": [""],
                "kontakt": [""],
                "Termín": [""]
            })
        # Kontingenční tabulka: řádky val_id, sloupce kategory, hodnoty počet
        if not df2.empty and "VAL_ID" in df2.columns and "Kategory" in df2.columns and "ROWCOUNT" in df2.columns:
            # Získat všechny kategorie z DB
            all_categories = sorted(set(kategory_list)) if kategory_list else ["nedefinováno"]
            pivot = pd.pivot_table(df2, index="VAL_ID", columns="Kategory", values="ROWCOUNT", aggfunc="sum", fill_value=0)
            # Přidat chybějící kategorie jako sloupce s nulou
            for cat in all_categories:
                if cat not in pivot.columns:
                    pivot[cat] = 0
            # Zachovat pořadí kategorií podle DB
            pivot = pivot[all_categories]
        else:
            pivot = pd.DataFrame()
        # Nejprve uložím pivot jako první list, pak ostatní
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            if not pivot.empty:
                pivot.to_excel(writer, sheet_name="pivot")
            df2.to_excel(writer, sheet_name="valOverview", index=False)
            df.to_excel(writer, sheet_name="valData", index=False)
            df_kategory.to_excel(writer, sheet_name="kategory", index=False)

        # Otevřít soubor pomocí openpyxl pro úpravy vzhledu
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, Alignment
        from openpyxl.cell.cell import Cell
        wb = load_workbook(filename)

        # Přidat nadpis na list kategory
        if "kategory" in wb.sheetnames:
            ws_kategory = wb["kategory"]
            ws_kategory.insert_rows(1)
            ws_kategory["A1"] = "Kategorie z tabulky SO081267.TMP_SST_VALIDATION_KATEGORY_CIS"
            ws_kategory["A1"].font = Font(bold=True, size=14)
            ws_kategory["A1"].alignment = Alignment(horizontal="center")
            max_col = ws_kategory.max_column
            ws_kategory.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

        # Vylepšení pivotu
        if "pivot" in wb.sheetnames:
            ws_pivot = wb["pivot"]
            # Přidat název tabulky nad pivot
            ws_pivot.insert_rows(1)
            ws_pivot["A1"] = "Přehled validací a jejich kategorií"
            ws_pivot["A1"].font = Font(bold=True, size=14)
            ws_pivot["A1"].alignment = Alignment(horizontal="center")
            # Sloučit buňky pro název přes všechny sloupce
            max_col = ws_pivot.max_column
            ws_pivot.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            # Tučné záhlaví
            for cell in ws_pivot[2]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Automatická šířka sloupců (ignorovat MergedCell)
            for col in ws_pivot.iter_cols(min_row=2):
                max_length = 0
                col_letter = col[0].column_letter if isinstance(col[0], Cell) else None
                if not col_letter:
                    continue
                for cell in col:
                    if isinstance(cell, Cell) and cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_pivot.column_dimensions[col_letter].width = max(12, max_length + 2)

        # Nastavit automatickou šířku sloupců na všech relevantních listech
        for sheet_name in ["valOverview", "valData", "kategory"]:
            if sheet_name in wb.sheetnames:
                ws_sheet = wb[sheet_name]
                for col in ws_sheet.iter_cols():
                    max_length = 0
                    col_letter = col[0].column_letter if isinstance(col[0], Cell) else None
                    if not col_letter:
                        continue
                    for cell in col:
                        if isinstance(cell, Cell) and cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws_sheet.column_dimensions[col_letter].width = max(12, max_length + 2)

        # Roletkové menu ve sloupci Kategory podle hodnot z listu kategory
        ws = wb["valOverview"]
        ws_kategory = wb["kategory"]
        last_row = ws_kategory.max_row
        for r in range(ws_kategory.max_row, 1, -1):
            if ws_kategory[f"A{r}"].value:
                last_row = r
                break
        formula_range = f"=kategory!$A$2:$A${last_row}"
        dv = DataValidation(type="list", formula1=formula_range, allow_blank=True)
        ws.add_data_validation(dv)
        for row in range(2, ws.max_row + 1):
            dv.add(ws[f"A{row}"])
        ws.auto_filter.ref = ws.dimensions

        wb.save(filename)
        print(f"Výsledek uložen do {filename}")
        cursor.close()
        connection.close()
    except Exception as e:
        print("Chyba při práci s databází:", e)
        connection = get_db_connection()
        cursor = connection.cursor()
        query = "SELECT * FROM K2_MIGUSER1.MSA_VAL_DESIGN ORDER BY VAL_ID"
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        data = cursor.fetchall()
        df = pd.DataFrame(data, columns=columns)
        output_dir = "vystupy"
        os.makedirs(output_dir, exist_ok=True)
        now_str = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"valDesign_{now_str}.xlsx")

        # Druhý list: vybrané sloupce a číslo z rowcount ve FINAL_QUERY
        cols_needed = ["VAL_ID", "NAME", "TABLE_NAME", "COLUMN_NAME", "SEVERITY", "STATUS", "FINAL_QUERY"]
        df2 = df[cols_needed].copy() if all(col in df.columns for col in cols_needed) else pd.DataFrame()
        # Načíst kategorie z DB
        kategory_dict = {}
        kategory_list = []
        try:
            cursor_kat = connection.cursor()
            cursor_kat.execute("SELECT VAL_ID, KATEGORY FROM SO081267.TMP_SST_VALIDATION_KATEGORY")
            for val_id, kategory in cursor_kat.fetchall():
                kategory_dict[val_id] = kategory
                kategory_list.append(kategory)
            cursor_kat.close()
        except Exception as e:
            print("Chyba při načítání kategorií z DB:", e)
        if not df2.empty:
            import re
            def extract_rowcount(val):
                if pd.isnull(val):
                    return None
                match = re.search(r"rowcount\s*=\s*(\d+)", str(val))
                return int(match.group(1)) if match else None
            def extract_sysdate(val):
                if pd.isnull(val):
                    return None
                import re
                match = re.search(r"sysdate\s*[=:]?\s*([0-9]{2})\.([0-9]{2})\.([0-9]{4})", str(val))
                if match:
                    return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
                match2 = re.search(r"([0-9]{4}-[0-9]{2}-[0-9]{2})", str(val))
                return match2.group(1) if match2 else None
            df2["DATUM_KONTROLY"] = df2["FINAL_QUERY"].apply(extract_sysdate)
            df2["ROWCOUNT"] = df2["FINAL_QUERY"].apply(extract_rowcount)
            df2 = df2.drop(columns=["FINAL_QUERY"])
            # Přidat sloupec Kategory podle DB, jinak 'nedefinováno'
            df2.insert(0, "Kategory", df2["VAL_ID"].map(lambda x: kategory_dict.get(x, "nedefinováno")))
            # Přesunout sloupec DATUM_KONTROLY před ROWCOUNT (vždy)
            cols = df2.columns.tolist()
            if "DATUM_KONTROLY" in cols and "ROWCOUNT" in cols:
                cols.remove("DATUM_KONTROLY")
                cols.insert(cols.index("ROWCOUNT"), "DATUM_KONTROLY")
                df2 = df2[cols]
        # List kategory: načíst hodnoty z tabulky SO081267.TMP_SST_VALIDATION_KATEGORY_CIS
        try:
            cursor_cis = connection.cursor()
            cursor_cis.execute("SELECT KATEGORY, ZODPOVEDNA_OSOBA, KONTAKT, TERMIN FROM SO081267.TMP_SST_VALIDATION_KATEGORY_CIS")
            data_cis = cursor_cis.fetchall()
            df_kategory = pd.DataFrame(data_cis, columns=["KATEGORY", "Zodpovědná osoba", "Kontakt", "Termín"])
            cursor_cis.close()
        except Exception as e:
            print("Chyba při načítání kategory cis z DB:", e)
            df_kategory = pd.DataFrame({
                "KATEGORIE": ["nedefinováno"],
                "Zodpovědná osoba": [""],
                "kontakt": [""],
                "Termín": [""]
            })
        # Kontingenční tabulka: řádky val_id, sloupce kategory, hodnoty počet
        if not df2.empty and "VAL_ID" in df2.columns and "Kategory" in df2.columns and "ROWCOUNT" in df2.columns:
            # Získat všechny kategorie z DB
            all_categories = sorted(set(kategory_list)) if kategory_list else ["nedefinováno"]
            pivot = pd.pivot_table(df2, index="VAL_ID", columns="Kategory", values="ROWCOUNT", aggfunc="sum", fill_value=0)
            # Přidat chybějící kategorie jako sloupce s nulou
            for cat in all_categories:
                if cat not in pivot.columns:
                    pivot[cat] = 0
            # Zachovat pořadí kategorií podle DB
            pivot = pivot[all_categories]
        else:
            pivot = pd.DataFrame()
        # Nejprve uložím pivot jako první list, pak ostatní
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            if not pivot.empty:
                pivot.to_excel(writer, sheet_name="pivot")
            df2.to_excel(writer, sheet_name="valOverview", index=False)
            df.to_excel(writer, sheet_name="valData", index=False)
            df_kategory.to_excel(writer, sheet_name="kategory", index=False)

        # Otevřít soubor pomocí openpyxl pro úpravy vzhledu
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, Alignment
        from openpyxl.cell.cell import Cell
        wb = load_workbook(filename)

        # Vylepšení pivotu
        if "pivot" in wb.sheetnames:
            ws_pivot = wb["pivot"]
            # Přidat název tabulky nad pivot
            ws_pivot.insert_rows(1)
            ws_pivot["A1"] = "Přehled validací a jejich řešitelů"
            ws_pivot["A1"].font = Font(bold=True, size=14)
            ws_pivot["A1"].alignment = Alignment(horizontal="center")
            # Sloučit buňky pro název přes všechny sloupce
            max_col = ws_pivot.max_column
            ws_pivot.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            # Tučné záhlaví
            for cell in ws_pivot[2]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Automatická šířka sloupců (ignorovat MergedCell)
            for col in ws_pivot.iter_cols(min_row=2):
                max_length = 0
                col_letter = col[0].column_letter if isinstance(col[0], Cell) else None
                if not col_letter:
                    continue
                for cell in col:
                    if isinstance(cell, Cell) and cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_pivot.column_dimensions[col_letter].width = max(12, max_length + 2)

        # Nastavit automatickou šířku sloupců na všech relevantních listech
        for sheet_name in ["valOverview", "valData", "kategory"]:
            if sheet_name in wb.sheetnames:
                ws_sheet = wb[sheet_name]
                for col in ws_sheet.iter_cols():
                    max_length = 0
                    col_letter = col[0].column_letter if isinstance(col[0], Cell) else None
                    if not col_letter:
                        continue
                    for cell in col:
                        if isinstance(cell, Cell) and cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws_sheet.column_dimensions[col_letter].width = max(12, max_length + 2)

        # Roletkové menu ve sloupci Kategory podle hodnot z listu kategory
        ws = wb["valOverview"]
        ws_kategory = wb["kategory"]
        last_row = ws_kategory.max_row
        for r in range(ws_kategory.max_row, 1, -1):
            if ws_kategory[f"A{r}"].value:
                last_row = r
                break
        formula_range = f"=kategory!$A$2:$A${last_row}"
        dv = DataValidation(type="list", formula1=formula_range, allow_blank=True)
        ws.add_data_validation(dv)
        for row in range(2, ws.max_row + 1):
            dv.add(ws[f"A{row}"])
        ws.auto_filter.ref = ws.dimensions

        wb.save(filename)
        print(f"Výsledek uložen do {filename}")
        cursor.close()
        connection.close()
    except Exception as e:
        print("Chyba při práci s databází:", e)

if __name__ == "__main__":
    main()
