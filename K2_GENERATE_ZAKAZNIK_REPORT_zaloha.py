"""
K2_GENERATE_ZAKAZNIK_REPORT.py
Základní struktura programu pro generování reportu zákazníků.
"""

from struktura_Dat.db_connect import get_db_connection
from datetime import datetime, timedelta
import os
from openpyxl.styles import PatternFill, Font, Alignment

def main():
    print("Skript spuštěn")
    print("Začínám vyhodnocení reportu...")
    print("Připojuji se k databázi...")
    connection = get_db_connection()
    cursor = connection.cursor()
    print("Připojeno.")

    # Definovat output_dir na začátku main()
    output_dir = os.path.join(os.path.dirname(__file__), 'vystupy')
    os.makedirs(output_dir, exist_ok=True)

    # Definovat now_str na začátku main()
    now_str = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Nastavit porovnání na dnešní datum a datum před týdnem
    date_today = datetime.now().date()
    date_compare = date_today - timedelta(days=7)

    # Počet záznamů se SEVERITY = 'ERRORS' pro dnešní datum
    print(f"Zjišťuji počet ERRORS pro {date_today}...")
    cursor.execute("SELECT COUNT(*) FROM SO081267.K2_VALIDACE_ZAKAZNIK WHERE UPPER(SEVERITY) = 'ERRORS' AND DATUMREPORTU = :1", [date_today])
    count_errors_today = cursor.fetchone()[0]
    print(f"Počet ERRORS pro {date_today}: {count_errors_today}")

    # Počet záznamů se SEVERITY = 'WARNINGS' pro dnešní datum
    print(f"Zjišťuji počet WARNINGS pro {date_today}...")
    cursor.execute("SELECT COUNT(*) FROM SO081267.K2_VALIDACE_ZAKAZNIK WHERE UPPER(SEVERITY) = 'WARNINGS' AND DATUMREPORTU = :1", [date_today])
    count_warnings_today = cursor.fetchone()[0]
    print(f"Počet WARNINGS pro {date_today}: {count_warnings_today}")

    # Počet záznamů se SEVERITY = 'ERRORS' pro porovnávací datum
    print(f"Zjišťuji počet ERRORS pro {date_compare}...")
    cursor.execute("SELECT COUNT(*) FROM SO081267.K2_VALIDACE_ZAKAZNIK WHERE UPPER(SEVERITY) = 'ERRORS' AND DATUMREPORTU = :1", [date_compare])
    count_errors_lastweek = cursor.fetchone()[0]
    print(f"Počet ERRORS pro {date_compare}: {count_errors_lastweek}")

    # Počet záznamů se SEVERITY = 'WARNINGS' pro porovnávací datum
    print(f"Zjišťuji počet WARNINGS pro {date_compare}...")
    cursor.execute("SELECT COUNT(*) FROM SO081267.K2_VALIDACE_ZAKAZNIK WHERE UPPER(SEVERITY) = 'WARNINGS' AND DATUMREPORTU = :1", [date_compare])
    count_warnings_lastweek = cursor.fetchone()[0]
    print(f"Počet WARNINGS pro {date_compare}: {count_warnings_lastweek}")

    print("\n--- Změny oproti minulému týdnu ---")
    error_diff = count_errors_today - count_errors_lastweek
    if error_diff < 0:
        if count_errors_lastweek == 0:
            print(f"Opraveno ERRORS: {abs(error_diff)} (minulý týden bylo 0)")
        else:
            print(f"Opraveno ERRORS: {abs(error_diff)} ({abs(error_diff)/count_errors_lastweek*100:.1f} %)")
    elif error_diff > 0:
        if count_errors_lastweek == 0:
            print(f"Přibylo ERRORS: {error_diff} (minulý týden bylo 0)")
        else:
            print(f"Přibylo ERRORS: {error_diff} (+{error_diff/count_errors_lastweek*100:.1f} %)")
    else:
        print(f"Počet ERRORS beze změny.")

    warning_diff = count_warnings_today - count_warnings_lastweek
    if warning_diff < 0:
        if count_warnings_lastweek == 0:
            print(f"Opraveno WARNINGS: {abs(warning_diff)} (minulý týden bylo 0)")
        else:
            print(f"Opraveno WARNINGS: {abs(warning_diff)} ({abs(warning_diff)/count_warnings_lastweek*100:.1f} %)")
    elif warning_diff > 0:
        if count_warnings_lastweek == 0:
            print(f"Přibylo WARNINGS: {warning_diff} (minulý týden bylo 0)")
        else:
            print(f"Přibylo WARNINGS: {warning_diff} (+{warning_diff/count_warnings_lastweek*100:.1f} %)")
    else:
        print(f"Počet WARNINGS beze změny.")

    # Export do Excelu
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Status dat"
    ws.append([f"Souhrn stavu čistoty dat za týden: {date_compare} - {date_today}"])
    ws.append([])
    ws.append(["Datum", "Počet ERRORS", "Počet WARNINGS"])
    ws.append([str(date_today), count_errors_today, count_warnings_today])
    ws.append([str(date_compare), count_errors_lastweek, count_warnings_lastweek])
    ws.append([])
    ws.append(["Změny oproti minulému týdnu:"])
    if error_diff < 0:
        if count_errors_lastweek == 0:
            ws.append([f"Opraveno ERRORS: {abs(error_diff)} (minulý týden bylo 0)"])
        else:
            ws.append([f"Opraveno ERRORS: {abs(error_diff)} ({abs(error_diff)/count_errors_lastweek*100:.1f} %)"])
    elif error_diff > 0:
        if count_errors_lastweek == 0:
            ws.append([f"Přibylo ERRORS: {error_diff} (minulý týden bylo 0)"])
        else:
            ws.append([f"Přibylo ERRORS: {error_diff} (+{error_diff/count_errors_lastweek*100:.1f} %)"])
    else:
        ws.append(["Počet ERRORS beze změny."])

    if warning_diff < 0:
        if count_warnings_lastweek == 0:
            ws.append([f"Opraveno WARNINGS: {abs(warning_diff)} (minulý týden bylo 0)"])
        else:
            ws.append([f"Opraveno WARNINGS: {abs(warning_diff)} ({abs(warning_diff)/count_warnings_lastweek*100:.1f} %)"])
    elif warning_diff > 0:
        if count_warnings_lastweek == 0:
            ws.append([f"Přibylo WARNINGS: {warning_diff} (minulý týden bylo 0)"])
        else:
            ws.append([f"Přibylo WARNINGS: {warning_diff} (+{warning_diff/count_warnings_lastweek*100:.1f} %)"])
    else:
        ws.append(["Počet WARNINGS beze změna."])

    # List TOP validace - automatické plnění z DB
    ws_top = wb.create_sheet(title="TOP validace")
    ws_top.append([f"TOP 10 nejčastějších validací k {date_today}"])
    ws_top.append([])
    ws_top.append(["VAL_ID", "SEVERITY", "DESCRIPTION", "Počet výskytů"])
    cursor.execute("""
        SELECT VAL_ID, SEVERITY, DESCRIPTION, COUNT(*) AS pocet
        FROM SO081267.K2_VALIDACE_ZAKAZNIK
        WHERE DATUMREPORTU = :1
        GROUP BY VAL_ID, SEVERITY, DESCRIPTION
        ORDER BY pocet DESC
    """, [date_today])
    top_validace = cursor.fetchall()
    for row in top_validace[:10]:
        ws_top.append(list(row))

    # List Nejproblematičtější validace (ručně vyplňované)
    ws_errors = wb.create_sheet(title="Nejproblematičtější validace")
    ws_errors.append(["Nejproblematičtější validace (pouze error validace)"])
    ws_errors.append([])
    ws_errors.append(["Validace", "Počet výskytů", "Blokace opravy", "Upřesnění/poznámka"])

    # List overviewValidations - přehled všech validací
    ws_overview = wb.create_sheet(title="overviewValidations")
    ws_overview.append([f"Přehled validací k {date_today}"])
    ws_overview.append([])
    ws_overview.append(["VAL_ID", "SEVERITY", "RESPONSIBLE", "DESCRIPTION", "Počet záznamů"])
    cursor.execute("""
        SELECT VAL_ID, SEVERITY, RESPONSIBLE, DESCRIPTION, COUNT(*) AS pocet
        FROM SO081267.K2_VALIDACE_ZAKAZNIK
        WHERE DATUMREPORTU = :1
        GROUP BY VAL_ID, SEVERITY, RESPONSIBLE, DESCRIPTION
        ORDER BY pocet DESC
    """, [date_today])
    overview_validace = cursor.fetchall()
    for row in overview_validace:
        ws_overview.append(list(row))

    # List Trend v čase - přejmenování a přesunutí na druhou pozici
    ws_trend = wb.create_sheet(title="Trend v čase", index=1)
    ws_trend.append(["Vývoj počtu chyb a oprav v čase (poslední 3 týdny, týdenní souhrn)"])
    ws_trend.append([])
    ws_trend.append(["Týden od", "Týden do", "Počet ERRORS", "Týdenní změna - ERRORS", "Počet WARNINGS", "Týdenní změna - WARNINGS"])
    # Získání dat za poslední 3 týdny (pondělí-neděle) + aktuální neúplný týden
    start_date = date_today - timedelta(days=date_today.weekday() + 21)
    prev_errors = None
    prev_warnings = None
    for i in range(3):
        week_start = start_date + timedelta(days=i*7)
        week_end = week_start + timedelta(days=6)
        cursor.execute("""
            SELECT SUM(CASE WHEN UPPER(SEVERITY) = 'ERRORS' THEN 1 ELSE 0 END) AS ERRORS,
                   SUM(CASE WHEN UPPER(SEVERITY) = 'WARNINGS' THEN 1 ELSE 0 END) AS WARNINGS
            FROM SO081267.K2_VALIDACE_ZAKAZNIK
            WHERE DATUMREPORTU >= :1 AND DATUMREPORTU <= :2
        """, [week_start, week_end])
        result = cursor.fetchone()
        errors = result[0] if result[0] is not None else 0
        warnings = result[1] if result[1] is not None else 0
        if i == 0:
            oprav_errors = ''
            oprav_warnings = ''
        else:
            oprav_errors = f"{errors - prev_errors:+d}"
            oprav_warnings = f"{warnings - prev_warnings:+d}"
        ws_trend.append([str(week_start), str(week_end), errors, oprav_errors, warnings, oprav_warnings])
        prev_errors = errors
        prev_warnings = warnings
    # Přidat aktuální neúplný týden
    week_start = date_today - timedelta(days=date_today.weekday())
    week_end = date_today
    cursor.execute("""
        SELECT SUM(CASE WHEN UPPER(SEVERITY) = 'ERRORS' THEN 1 ELSE 0 END) AS ERRORS,
               SUM(CASE WHEN UPPER(SEVERITY) = 'WARNINGS' THEN 1 ELSE 0 END) AS WARNINGS
        FROM SO081267.K2_VALIDACE_ZAKAZNIK
        WHERE DATUMREPORTU >= :1 AND DATUMREPORTU <= :2
    """, [week_start, week_end])
    result = cursor.fetchone()
    errors = result[0] if result[0] is not None else 0
    warnings = result[1] if result[1] is not None else 0
    oprav_errors = f"{errors - prev_errors:+d}" if prev_errors is not None else ''
    oprav_warnings = f"{warnings - prev_warnings:+d}" if prev_warnings is not None else ''
    ws_trend.append([str(week_start), str(week_end), errors, oprav_errors, warnings, oprav_warnings])

    # Po naplnění dat do ws_trend přidat graf vývoje ERRORS a WARNINGS
    # Graf s dvěma čarami: ERRORS a WARNINGS, bez týdenní změny
    import matplotlib.pyplot as plt
    weeks = []
    errors = []
    warnings = []
    # Přidat týden 11-17.8 ručně na začátek
    weeks.append("2025-08-11 - 2025-08-17")
    errors.append(0)  # Pokud znáte hodnotu, upravte
    warnings.append(0)  # Pokud znáte hodnotu, upravte
    for row in ws_trend.iter_rows(min_row=5, max_row=ws_trend.max_row, values_only=True):
        weeks.append(f"{row[0]} - {row[1]}")
        errors.append(row[2])
        warnings.append(row[4])
    plt.figure(figsize=(10,6))
    plt.plot(weeks, errors, marker='o', label='ERRORS')
    plt.plot(weeks, warnings, marker='o', label='WARNINGS')
    plt.title('Vývoj chyb a varování v čase (poslední 3 týdny)')
    plt.xlabel('Týden')
    plt.ylabel('Počet')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    # Nastavit popisky osy Y po 50
    import matplotlib.ticker as ticker
    ax = plt.gca()
    ax.yaxis.set_major_locator(ticker.MultipleLocator(50))
    # Přidat hodnoty k bodům v grafu
    for i, val in enumerate(errors):
        plt.text(i, val, str(val), ha='center', va='bottom', fontsize=9, color='blue')
    for i, val in enumerate(warnings):
        plt.text(i, val, str(val), ha='center', va='bottom', fontsize=9, color='orange')
    img_path = os.path.join(output_dir, f'trend_v_case_{now_str}.png')
    plt.savefig(img_path)
    print(f'Graf byl uložen jako {img_path}')

    excel_path = os.path.join(output_dir, f'K2_status_cistota_dat_{now_str}.xlsx')
    img_path = os.path.join(output_dir, f'trend_v_case_{now_str}.png')
    wb.save(excel_path)
    print(f"Výsledky byly uloženy do {excel_path}")

    # Obarvení záhlaví a nadpisů na všech listech
    from openpyxl.styles import PatternFill, Font, Alignment
    title_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
    for ws in wb.worksheets:
        # Nadpis (první řádek)
        for cell in ws[1]:
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = header_align
        # Záhlaví (najít první ne-prázdný řádek s více než 1 buňkou)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if len([c for c in row if c.value]) > 1:
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                break
    wb.save(excel_path)
    print(f"Výsledky byly uloženy do {excel_path}")

    # Vložení obrázku grafu pod tabulku na list Trend v čase
    from openpyxl.drawing.image import Image as XLImage
    img_path = os.path.join(output_dir, f'trend_v_case_{now_str}.png')
    if os.path.exists(img_path):
        img = XLImage(img_path)
        # Najít první volný řádek pod tabulkou
        last_row = ws_trend.max_row + 2
        ws_trend.add_image(img, f"A{last_row}")
    wb.save(excel_path)
    print(f"Výsledky byly uloženy do {excel_path}")

    # Automatické nastavení šířky sloupců podle obsahu na všech listech
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"Výsledky byly uloženy do {excel_path}")

    # Trend v čase: každý den zvlášť, ne součet za týden
    ws_trend = wb.create_sheet(title="Trend v čase (denní)", index=2)
    ws_trend.append(["Datum", "Počet ERRORS", "Počet WARNINGS"])
    # Získat data za poslední 3 týdny + aktuální den
    start_date = date_today - timedelta(days=20)
    for i in range(21):
        day = start_date + timedelta(days=i)
        cursor.execute("""
            SELECT COUNT(*) FROM SO081267.K2_VALIDACE_ZAKAZNIK WHERE UPPER(SEVERITY) = 'ERRORS' AND DATUMREPORTU = :1
        """, [day])
        errors = cursor.fetchone()[0]
        cursor.execute("""
            SELECT COUNT(*) FROM SO081267.K2_VALIDACE_ZAKAZNIK WHERE UPPER(SEVERITY) = 'WARNINGS' AND DATUMREPORTU = :1
        """, [day])
        warnings = cursor.fetchone()[0]
        ws_trend.append([str(day), errors, warnings])

    # Graf ERRORS a WARNINGS za poslední 3 týdny (denní data)
    #img_path = os.path.join(output_dir, f'trend_v_case_denni_{now_str}.png')
    #plt.figure(figsize=(10,6))
    #plt.plot([row[0] for row in ws_trend.iter_rows(min_row=2, max_row=ws_trend.max_row, values_only=True)], [row[1] for row in ws_trend.iter_rows(min_row=2, max_row=ws_trend.max_row, values_only=True)], marker='o', label='ERRORS')
    #plt.plot([row[0] for row in ws_trend.iter_rows(min_row=2, max_row=ws_trend.max_row, values_only=True)], [row[2] for row in ws_trend.iter_rows(min_row=2, max_row=ws_trend.max_row, values_only=True)], marker='o', label='WARNINGS')
    #plt.title('Vývoj chyb a varování v čase (poslední 3 týdny, denní souhrn)')
    #plt.xlabel('Datum')
    #plt.ylabel('Počet')
    #plt.legend()
    #plt.grid(True)
    #plt.tight_layout()
    # Nastavit popisky osy Y po 50
    #ax = plt.gca()
    #ax.yaxis.set_major_locator(ticker.MultipleLocator(50))
    # Přidat hodnoty k bodům v grafu
    #for i, val in enumerate([row[1] for row in ws_trend.iter_rows(min_row=2, max_row=ws_trend.max_row, values_only=True)]):
    #    plt.text(i, val, str(val), ha='center', va='bottom', fontsize=9, color='blue')
    #for i, val in enumerate([row[2] for row in ws_trend.iter_rows(min_row=2, max_row=ws_trend.max_row, values_only=True)]):
    #    plt.text(i, val, str(val), ha='center', va='bottom', fontsize=9, color='orange')
    #plt.xticks(rotation=45)
    #plt.subplots_adjust(bottom=0.2)
    #plt.savefig(img_path)
    #print(f'Graf (denní) byl uložen jako {img_path}')

    # Vložení obrázku grafu pod tabulku na list Trend v čase (denní)
    #img_path = os.path.join(output_dir, f'trend_v_case_denni_{now_str}.png')
    #if os.path.exists(img_path):
    #    img = XLImage(img_path)
        # Najít první volný řádek pod tabulkou
    #    last_row = ws_trend.max_row + 2
    #    ws_trend.add_image(img, f"A{last_row}")
    #wb.save(excel_path)
    #print(f"Výsledky byly uloženy do {excel_path}")

    cursor.close()
    connection.close()

if __name__ == "__main__":
    main()